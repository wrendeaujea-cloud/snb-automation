import shutil
from pathlib import Path
import pandas as pd

from datetime import datetime
from zoneinfo import ZoneInfo 

#to open again and adjust font
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# ==========CONFIG===========
# destination folder inside workflow
DEST_FOLDER = Path("output")
DEST_FOLDER.mkdir(exist_ok=True)

# OneDrive direct links
ContactReport_URL = "https://lexmarkad-my.sharepoint.com/:x:/r/personal/wrendeaujea_berame_lexmark_com/Documents/Desktop/SNB/AM0165SX_New%20-%20Contacts%20For%20Account%20with%20Row%20Id.xlsx?d=wa725262370634c2784fbd96941597db0&csf=1&web=1&e=fzNcA8"
MP8032Report_URL = "https://lexmarkad-my.sharepoint.com/:x:/r/personal/wrendeaujea_berame_lexmark_com/Documents/Desktop/SNB/MP8032SP%20-%20Asset%20Detail%20Report%20-%20Extended.xlsx?d=w246fc6a813a64d47bc04c2316988a550&csf=1&web=1&e=0jWKiX"

# local filenames in workflow
ContactReport_FILE = DEST_FOLDER / "AM0165SX_New - Contacts For Account with Row Id.xlsx"
MP8032Report_FILE = DEST_FOLDER / "MP8032SP - Asset Detail Report - Extended.xlsx"

#==================

# download Contact report
r = requests.get(ContactReport_URL)
r.raise_for_status()
with open(ContactReport_FILE, "wb") as f:
    f.write(r.content)

# download Asset report
r = requests.get(MP8032Report_URL)
r.raise_for_status()
with open(MP8032Report_FILE, "wb") as f:
    f.write(r.content)

print("Files downloaded successfully")

def main():
    #create destination folder if it doesn't exist
    DEST_FOLDER.mkdir(parents=True, exist_ok=True)

    #define full file path
    src_1 = SOURCE_FOLDER / ContactReport
    src_2 = SOURCE_FOLDER / MP8032Report

    dest_1 = DEST_FOLDER / ContactReport
    dest_2 = DEST_FOLDER / MP8032Report

    #copy files
    shutil.copy(src_1, dest_1)
    shutil.copy(src_2, dest_2)

    print("Files copied successfully")

if __name__ == "__main__":
    main()

#load excel files
contact_df = pd.read_excel(DEST_FOLDER / ContactReport, header = 1)
asset_df = pd.read_excel(DEST_FOLDER / MP8032Report, header = 1)

#merge (left join)
merge_df = asset_df.merge(
    contact_df[["#Contact.ContactUID", "Contact.WorkPhone"]],
    left_on="Contact User ID",
    right_on="#Contact.ContactUID",
    how = "left"
)

#drop redudant lookup column
merge_df = merge_df.drop(columns = ["#Contact.ContactUID"])

#move workphone column
phone_col = merge_df.pop("Contact.WorkPhone") #remove temporarily
merge_df.insert(23, "Phone#", phone_col) #insert it back after email column

#columns to drop
columns_to_drop = [
    "CHL4", "CHL5", "CHL6", "CHL7", "CHL8",
    "Network Name", "Computer Name", "Network Topology", "Device Tag ServiceTag", "Device Tag Service Partner", "Special Usage", "Stored Date", "Project", "Equipment ID"
]
merge_df = merge_df.drop(columns = columns_to_drop, errors="ignore")

#current date in Eastern Time
eastern = ZoneInfo("America/New_York")
today_str = datetime.now(tz=eastern).strftime("%Y-%m-%d")

#save to excel
output_file = DEST_FOLDER / f"SNB Asset Detail Report {today_str}.xlsx"
merge_df.to_excel(output_file, index = False)
print(f"Merged report saved to: {output_file}")

#load the file padas just saved
wb = load_workbook(output_file)
ws = wb.active

#remove headers borders for row only
no_border = Border(
    left=Side(border_style=None),
    right=Side(border_style=None),
    top=Side(border_style=None),
    bottom=Side(border_style=None)
)
for cell in ws[1]: #header row
    cell.border = no_border

#insert new top row
ws.insert_rows(1)
ws["A1"] = "#LXK UP MPS Asset"
ws["B1"] = "MP8032SP- Asset Detail - Conversion"

#apply arial 10 default excel font
arial10 = Font(name="ARIAL", size = 10)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = arial10

#save again
wb.save(output_file)

print("Fonts and border applied successfully")
